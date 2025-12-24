import os
import tempfile
import platform
import traceback

# Import libraries with safety
try:
    from PyPDF2 import PdfReader
    from pdf2docx import Converter
    from docx import Document
    from PIL import Image
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError as e:
    print(f"⚠️ Missing library: {e}")
    # Consider adding a logger here for production

class FileConverter:
    CONVERTERS = {}
    
    @classmethod
    def register_converter(cls, conversion_type):
        def decorator(func):
            cls.CONVERTERS[conversion_type] = func
            return func
        return decorator
    
    @staticmethod
    def _create_temp_output(extension=''):
        """Create temp file path safely without keeping the handle open"""
        suffix = extension if extension.startswith('.') else f'.{extension}'
        # Using mkstemp and closing the handle immediately is the safest way
        # to ensure other libraries can open the file for writing.
        fd, path = tempfile.mkstemp(suffix=suffix)
        os.close(fd) 
        return path

    @classmethod
    def convert_file(cls, conversion_type, input_path):
        """Main conversion engine"""
        print(f"--- Processing: {conversion_type} ---")
        converter = cls.CONVERTERS.get(conversion_type)
        if not converter:
            return False, None, f"Unsupported conversion: {conversion_type}"
        
        if not os.path.exists(input_path):
            return False, None, "Source file not found on server."

        # Map extensions
        ext_map = {
            'pdf_to_word': '.docx', 
            'word_to_pdf': '.pdf',
            'pdf_to_excel': '.xlsx', 
            'excel_to_pdf': '.pdf',
            'pdf_to_image': '.png', 
            'image_to_pdf': '.pdf',
        }
        
        output_path = cls._create_temp_output(ext_map.get(conversion_type, '.tmp'))
        
        try:
            success, message = converter(input_path, output_path)
            
            if success and os.path.exists(output_path) and os.path.getsize(output_path) > 0:
                with open(output_path, 'rb') as f:
                    content = f.read()
                
                # Cleanup temp file after reading into memory
                if os.path.exists(output_path):
                    os.unlink(output_path)
                return True, content, message
            else:
                if os.path.exists(output_path): os.unlink(output_path)
                return False, None, message or "Conversion produced an empty file."
                
        except Exception as e:
            if os.path.exists(output_path): os.unlink(output_path)
            traceback.print_exc()
            return False, None, f"Converter Internal Error: {str(e)}"

# ============================================================================
# INDIVIDUAL CONVERTERS
# ============================================================================

@FileConverter.register_converter('pdf_to_word')
def pdf_to_word(input_path, output_path):
    try:
        cv = Converter(input_path)
        cv.convert(output_path, start=0, end=None)
        cv.close()
        return True, "PDF to Word completed successfully."
    except Exception as e:
        return False, f"pdf2docx error: {str(e)}"

@FileConverter.register_converter('word_to_pdf')
def word_to_pdf(input_path, output_path):
    """Headless-safe Word to PDF with fallback"""
    # 1. Primary Method: docx2pdf (Requires Word installed on local machine)
    try:
        from docx2pdf import convert
        # On Windows/Mac with Word installed, this is the best quality
        convert(input_path, output_path)
        return True, "Success via docx2pdf"
    except Exception:
        # 2. Fallback: Manual ReportLab (Works on Linux/Servers without Word)
        return manual_word_to_pdf_logic(input_path, output_path)

def manual_word_to_pdf_logic(input_path, output_path):
    try:
        doc = Document(input_path)
        pdf_doc = SimpleDocTemplate(output_path, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        for para in doc.paragraphs:
            if para.text.strip():
                # Encoding cleanup to prevent ReportLab crashes on special characters
                clean_text = para.text.encode('utf-8', 'ignore').decode('utf-8')
                story.append(Paragraph(clean_text, styles['Normal']))
                story.append(Spacer(1, 0.1*inch))
        pdf_doc.build(story)
        return True, "Success (Manual Fallback)"
    except Exception as e:
        return False, f"Manual conversion failed: {str(e)}"

@FileConverter.register_converter('pdf_to_excel')
def pdf_to_excel(input_path, output_path):
    try:
        reader = PdfReader(input_path)
        wb = openpyxl.Workbook()
        ws = wb.active
        for page in reader.pages:
            text = page.extract_text()
            if text:
                for line in text.split('\n'):
                    if line.strip():
                        # Simple split by space; for better results, use libraries like 'tabula-py'
                        ws.append(line.split())
        wb.save(output_path)
        return True, "PDF to Excel completed."
    except Exception as e:
        return False, str(e)

@FileConverter.register_converter('pdf_to_image')
def pdf_to_image(input_path, output_path):
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(input_path)
        page = doc[0]  # Extracts the first page
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2)) # 2x zoom for better quality
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        img.save(output_path, "PNG")
        doc.close()
        return True, "PDF to Image completed."
    except Exception as e:
        return False, f"PyMuPDF error: {str(e)}"

@FileConverter.register_converter('image_to_pdf')
def image_to_pdf(input_path, output_path):
    try:
        img = Image.open(input_path).convert('RGB')
        # Maintain high resolution
        img.save(output_path, "PDF", resolution=100.0)
        return True, "Image to PDF completed."
    except Exception as e:
        return False, str(e)

@FileConverter.register_converter('excel_to_pdf')
def excel_to_pdf(input_path, output_path):
    try:
        wb = openpyxl.load_workbook(input_path, data_only=True)
        sheet = wb.active
        pdf_doc = SimpleDocTemplate(output_path, pagesize=A4)
        data = []
        for row in sheet.iter_rows(values_only=True):
            # Limit cell content length to avoid table overflow
            data.append([str(c)[:50] if c is not None else "" for c in row])
        
        if not data:
            return False, "Excel sheet is empty."

        table = Table(data)
        table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ]))
        pdf_doc.build([table])
        return True, "Excel to PDF completed."
    except Exception as e:
        return False, str(e)